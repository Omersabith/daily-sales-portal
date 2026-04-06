from __future__ import annotations

import csv
import io
import os
from datetime import date, datetime
from functools import wraps
from pathlib import Path

from db_store import clear_import_preview, ensure_database, load_all_data, load_import_preview, save_all_data, save_import_preview
from flask import Flask, flash, redirect, render_template, request, session, url_for, Response
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = Path(os.environ.get("APP_DATA_DIR", BASE_DIR))
DATA_FILE = Path(os.environ.get("DATA_FILE", str(DEFAULT_DATA_DIR / "daily_sales_portal.xlsx")))
DATABASE = os.environ.get("DATABASE_URL", str(Path(os.environ.get("DATABASE_PATH", str(DEFAULT_DATA_DIR / "sales_portal.db")))))

USERS_SHEET = "Users"
SKUS_SHEET = "SKUs"
TARGETS_SHEET = "Targets"
SALES_SHEET = "Sales"
AUDIT_SHEET = "Audit Log"
LEGACY_SALES_SHEET = "Daily Sales"

USERS_COLUMNS = ["username", "password_hash", "full_name", "role", "location", "active"]
SKUS_COLUMNS = ["sku_code", "sku_name", "category", "default_price", "active"]
TARGETS_COLUMNS = ["id", "target_from", "target_to", "promoter_username", "promoter_name", "target_amount", "notes", "updated_at"]
SALES_COLUMNS = [
    "id",
    "sale_date",
    "username",
    "promoter_name",
    "location",
    "sku_code",
    "sku_name",
    "category",
    "quantity",
    "selling_price",
    "sale_amount",
    "notes",
    "created_at",
    "updated_at",
]
AUDIT_COLUMNS = ["id", "event_time", "actor_username", "actor_role", "action", "entity_type", "entity_id", "details"]

DEFAULT_USERS = [
    {
        "username": "admin",
        "password_hash": generate_password_hash("admin123"),
        "full_name": "Portal Admin",
        "role": "super_admin",
        "location": "Head Office",
        "active": "yes",
    },
    {
        "username": "promoter1",
        "password_hash": generate_password_hash("promoter123"),
        "full_name": "Promoter One",
        "role": "promoter",
        "location": "Dubai Mall",
        "active": "yes",
    },
]

DEFAULT_SKUS = [
    {
        "sku_code": "A1001",
        "sku_name": "Anker Charger 20W",
        "category": "Chargers",
        "default_price": 79.0,
        "active": "yes",
    },
    {
        "sku_code": "A2001",
        "sku_name": "Anker Power Bank 10000mAh",
        "category": "Power Banks",
        "default_price": 149.0,
        "active": "yes",
    },
    {
        "sku_code": "A3001",
        "sku_name": "Anker USB-C Cable",
        "category": "Cables",
        "default_price": 39.0,
        "active": "yes",
    },
]


app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "daily-sales-portal-secret")
app.config["DATA_FILE"] = str(DATA_FILE)
app.config["DATABASE"] = DATABASE


def to_int(value: object) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def to_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def normalize_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_date_value(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text_value = normalize_text(value)
    if not text_value:
        return ""

    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text_value, fmt).date().isoformat()
        except ValueError:
            continue
    return text_value


def read_sheet(workbook, sheet_name: str, columns: list[str]) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [normalize_text(cell) for cell in rows[0]]
    index_map = {column: idx for idx, column in enumerate(header)}
    records: list[dict[str, object]] = []

    for row in rows[1:]:
        if row is None or not any(cell is not None and str(cell).strip() for cell in row):
            continue

        record: dict[str, object] = {}
        for column in columns:
            value = row[index_map[column]] if column in index_map and index_map[column] < len(row) else ""
            record[column] = "" if value is None else value
        records.append(record)

    return records


def read_generic_sheet(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [normalize_text(cell) for cell in rows[0]]
    records: list[dict[str, object]] = []
    for row in rows[1:]:
        if row is None or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record: dict[str, object] = {}
        for index, column in enumerate(header):
            value = row[index] if index < len(row) else ""
            record[column] = "" if value is None else value
        records.append(record)
    return records


def write_sheet(workbook, sheet_name: str, columns: list[str], records: list[dict[str, object]]) -> None:
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)

    sheet = workbook.create_sheet(sheet_name)
    sheet.append(columns)
    for record in records:
        sheet.append([record.get(column, "") for column in columns])


def get_workbook():
    return load_workbook(app.config["DATA_FILE"])


def save_workbook(workbook) -> None:
    workbook.save(app.config["DATA_FILE"])
    workbook.close()


def initialize_workbook() -> None:
    if not str(app.config["DATABASE"]).startswith(("postgres://", "postgresql://")):
        Path(app.config["DATABASE"]).parent.mkdir(parents=True, exist_ok=True)
    Path(app.config["DATA_FILE"]).parent.mkdir(parents=True, exist_ok=True)
    ensure_database(app.config["DATABASE"], app.config["DATA_FILE"])


initialize_workbook()


def get_all_data() -> dict[str, list[dict[str, object]]]:
    return load_all_data(app.config["DATABASE"])


def write_all_data(data: dict[str, list[dict[str, object]]]) -> None:
    save_all_data(app.config["DATABASE"], data)


def get_user(username: str) -> dict[str, object] | None:
    for user in get_all_data()["users"]:
        if normalize_text(user.get("username")).lower() == username.lower():
            user["username"] = normalize_text(user.get("username"))
            user["full_name"] = normalize_text(user.get("full_name"))
            user["role"] = normalize_text(user.get("role"))
            user["location"] = normalize_text(user.get("location"))
            user["active"] = normalize_text(user.get("active")).lower()
            return user
    return None


def get_active_skus() -> list[dict[str, object]]:
    skus = []
    for sku in get_all_data()["skus"]:
        if normalize_text(sku.get("active")).lower() != "yes":
            continue
        skus.append(
            {
                "sku_code": normalize_text(sku.get("sku_code")),
                "sku_name": normalize_text(sku.get("sku_name")),
                "category": normalize_text(sku.get("category")),
                "default_price": to_float(sku.get("default_price")),
            }
        )
    return sorted(skus, key=lambda item: (item["category"].lower(), item["sku_name"].lower()))


def get_sku_map() -> dict[str, dict[str, object]]:
    return {sku["sku_code"]: sku for sku in get_active_skus()}


def next_id(records: list[dict[str, object]]) -> int:
    return max((to_int(record.get("id")) for record in records), default=0) + 1


def current_user() -> dict[str, str] | None:
    username = normalize_text(session.get("username"))
    if not username:
        return None
    return {
        "username": username,
        "full_name": normalize_text(session.get("full_name")),
        "role": normalize_text(session.get("role")),
        "location": normalize_text(session.get("location")),
    }


def is_admin_role(role: str) -> bool:
    return role in {"super_admin", "admin"}


def is_super_admin(user: dict[str, str] | None) -> bool:
    return user is not None and user["role"] == "super_admin"


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if current_user() is None:
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        user = current_user()
        if user is None:
            return redirect(url_for("login"))
        if not is_admin_role(user["role"]):
            flash("Admin access is required.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped_view


def super_admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        user = current_user()
        if user is None:
            return redirect(url_for("login"))
        if not is_super_admin(user):
            flash("Super admin access is required.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped_view


@app.context_processor
def inject_globals():
    return {"current_user": current_user(), "today_iso": date.today().isoformat()}


def append_audit_log(
    data: dict[str, list[dict[str, object]]],
    action: str,
    entity_type: str,
    entity_id: object,
    details: str,
) -> None:
    user = current_user() or {"username": "system", "role": "system"}
    data["audit_logs"].append(
        {
            "id": next_id(data["audit_logs"]),
            "event_time": datetime.now().isoformat(timespec="seconds"),
            "actor_username": user["username"],
            "actor_role": user["role"],
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "details": details,
        }
    )


def build_row_result(row_number: int, status: str, messages: list[str], preview: dict[str, object] | None = None) -> dict[str, object]:
    return {
        "row_number": row_number,
        "status": status,
        "messages": messages,
        "preview": preview or {},
    }


def parse_decimal_import_value(raw_value: object, field_name: str, allow_zero: bool = True) -> tuple[float, list[str]]:
    text_value = normalize_text(raw_value)
    if not text_value:
        return 0.0, [f"{field_name} is required."]
    try:
        parsed_value = float(text_value)
    except ValueError:
        return 0.0, [f"{field_name} must be a number."]
    if parsed_value < 0 or (not allow_zero and parsed_value == 0):
        return 0.0, [f"{field_name} must be greater than or equal to zero." if allow_zero else f"{field_name} must be greater than zero."]
    return round(parsed_value, 2), []


def import_backend_template(file_storage) -> dict[str, object]:
    errors: list[str] = []
    users: list[dict[str, object]] = []
    skus: list[dict[str, object]] = []
    targets: list[dict[str, object]] = []
    historical_sales: list[dict[str, object]] = []
    user_rows: list[dict[str, object]] = []
    sku_rows: list[dict[str, object]] = []
    target_rows: list[dict[str, object]] = []
    historical_rows: list[dict[str, object]] = []

    try:
        workbook = load_workbook(file_storage)
    except Exception:
        return {
            "filename": getattr(file_storage, "filename", "uploaded-workbook.xlsx"),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "is_valid": False,
            "errors": ["Uploaded file could not be read as an Excel workbook."],
            "summary": {},
            "row_results": {"users": [], "skus": [], "targets": [], "historical_sales": []},
            "imported_data": None,
        }

    required_sheets = {"Users", "SKUs", "Targets", "Historical Sales"}
    missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
    if missing_sheets:
        workbook.close()
        return {
            "filename": getattr(file_storage, "filename", "uploaded-workbook.xlsx"),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "is_valid": False,
            "errors": [f"Missing sheet(s): {', '.join(missing_sheets)}."],
            "summary": {},
            "row_results": {"users": [], "skus": [], "targets": [], "historical_sales": []},
            "imported_data": None,
        }

    imported_users = read_generic_sheet(workbook, "Users")
    imported_skus = read_generic_sheet(workbook, "SKUs")
    imported_targets = read_generic_sheet(workbook, "Targets")
    imported_historical_sales = read_generic_sheet(workbook, "Historical Sales")
    workbook.close()

    seen_usernames: set[str] = set()
    for index, row in enumerate(imported_users, start=2):
        row_errors: list[str] = []
        username = normalize_text(row.get("username")).lower()
        plain_password = normalize_text(row.get("plain_password") or row.get("password"))
        password_hash = normalize_text(row.get("password_hash"))
        full_name = normalize_text(row.get("full_name"))
        role = normalize_text(row.get("role")).lower()
        location = normalize_text(row.get("location"))
        active_raw = normalize_text(row.get("active")).lower()
        active = "yes" if active_raw == "yes" else "no"

        if not username:
            row_errors.append("username is required.")
        if username and username in seen_usernames:
            row_errors.append(f"username '{username}' is duplicated.")
        if role not in {"super_admin", "admin", "promoter"}:
            row_errors.append("role must be super_admin, admin, or promoter.")
        if not full_name:
            row_errors.append("full_name is required.")
        if not location:
            row_errors.append("location is required.")
        if not plain_password and not password_hash:
            row_errors.append("provide plain_password or password_hash.")
        if active_raw not in {"", "yes", "no"}:
            row_errors.append("active must be yes or no.")

        preview_record = {
            "username": username,
            "full_name": full_name,
            "role": role,
            "location": location,
            "active": active,
            "password_source": "plain_password" if plain_password else "password_hash",
        }
        if row_errors:
            errors.extend([f"Users row {index}: {message}" for message in row_errors])
            user_rows.append(build_row_result(index, "invalid", row_errors, preview_record))
            continue

        user_record = {
            "username": username,
            "password_hash": generate_password_hash(plain_password) if plain_password else password_hash,
            "full_name": full_name,
            "role": role,
            "location": location,
            "active": active,
        }
        users.append(user_record)
        seen_usernames.add(username)
        user_rows.append(build_row_result(index, "valid", ["Ready to import."], preview_record))

    seen_skus: set[str] = set()
    for index, row in enumerate(imported_skus, start=2):
        row_errors: list[str] = []
        sku_code = normalize_text(row.get("sku_code")).upper()
        sku_name = normalize_text(row.get("sku_name"))
        category = normalize_text(row.get("category"))
        default_price_raw = row.get("default_price")
        active_raw = normalize_text(row.get("active")).lower()
        active = "yes" if active_raw == "yes" else "no"

        if not sku_code:
            row_errors.append("sku_code is required.")
        if sku_code and sku_code in seen_skus:
            row_errors.append(f"sku_code '{sku_code}' is duplicated.")
        if not sku_name:
            row_errors.append("sku_name is required.")
        if not category:
            row_errors.append("category is required.")
        default_price, price_errors = parse_decimal_import_value(default_price_raw, "default_price")
        row_errors.extend(price_errors)
        if active_raw not in {"", "yes", "no"}:
            row_errors.append("active must be yes or no.")

        preview_record = {
            "sku_code": sku_code,
            "sku_name": sku_name,
            "category": category,
            "default_price": default_price,
            "active": active,
        }
        if row_errors:
            errors.extend([f"SKUs row {index}: {message}" for message in row_errors])
            sku_rows.append(build_row_result(index, "invalid", row_errors, preview_record))
            continue

        sku_record = {
            "sku_code": sku_code,
            "sku_name": sku_name,
            "category": category,
            "default_price": default_price,
            "active": active,
        }
        skus.append(sku_record)
        seen_skus.add(sku_code)
        sku_rows.append(build_row_result(index, "valid", ["Ready to import."], preview_record))

    promoter_lookup = {
        normalize_text(user["username"]).lower(): normalize_text(user["full_name"])
        for user in users
        if normalize_text(user["role"]).lower() == "promoter" and normalize_text(user["active"]).lower() == "yes"
    }
    next_target_id = 1
    for index, row in enumerate(imported_targets, start=2):
        row_errors: list[str] = []
        target_from = normalize_date_value(row.get("target_from") or row.get("target_date"))
        target_to = normalize_date_value(row.get("target_to") or row.get("target_date"))
        promoter_username = normalize_text(row.get("promoter_username") or row.get("username")).lower()
        target_amount, amount_errors = parse_decimal_import_value(row.get("target_amount"), "target_amount")
        notes = normalize_text(row.get("notes"))
        updated_at = normalize_text(row.get("updated_at")) or datetime.now().isoformat(timespec="seconds")

        if not target_from:
            row_errors.append("target_from is required.")
            from_date = None
        else:
            try:
                from_date = datetime.strptime(target_from, "%Y-%m-%d")
            except ValueError:
                row_errors.append("target_from must use YYYY-MM-DD format.")
                from_date = None
        if not target_to:
            row_errors.append("target_to is required.")
            to_date = None
        else:
            try:
                to_date = datetime.strptime(target_to, "%Y-%m-%d")
            except ValueError:
                row_errors.append("target_to must use YYYY-MM-DD format.")
                to_date = None
        if from_date is not None and to_date is not None and to_date < from_date:
            row_errors.append("target_to cannot be earlier than target_from.")
        if promoter_username not in promoter_lookup:
            row_errors.append(f"promoter_username '{promoter_username}' must exist as an active promoter in Users sheet.")
        row_errors.extend(amount_errors)

        preview_record = {
            "target_from": target_from,
            "target_to": target_to,
            "promoter_username": promoter_username,
            "promoter_name": promoter_lookup.get(promoter_username, ""),
            "target_amount": target_amount,
        }
        if row_errors:
            errors.extend([f"Targets row {index}: {message}" for message in row_errors])
            target_rows.append(build_row_result(index, "invalid", row_errors, preview_record))
            continue

        target_record = {
            "id": next_target_id,
            "target_from": target_from,
            "target_to": target_to,
            "promoter_username": promoter_username,
            "promoter_name": promoter_lookup[promoter_username],
            "target_amount": target_amount,
            "notes": notes,
            "updated_at": updated_at,
        }
        targets.append(target_record)
        target_rows.append(build_row_result(index, "valid", ["Ready to import."], preview_record))
        next_target_id += 1

    next_history_id = 1
    for index, row in enumerate(imported_historical_sales, start=2):
        row_errors: list[str] = []
        period_from = normalize_date_value(row.get("period_from"))
        period_to = normalize_date_value(row.get("period_to"))
        promoter_username = normalize_text(row.get("promoter_username") or row.get("username")).lower()
        total_sales, amount_errors = parse_decimal_import_value(row.get("total_sales"), "total_sales")
        notes = normalize_text(row.get("notes"))
        updated_at = normalize_text(row.get("updated_at")) or datetime.now().isoformat(timespec="seconds")

        if not period_from:
            row_errors.append("period_from is required.")
            from_date = None
        else:
            try:
                from_date = datetime.strptime(period_from, "%Y-%m-%d")
            except ValueError:
                row_errors.append("period_from must use YYYY-MM-DD format.")
                from_date = None
        if not period_to:
            row_errors.append("period_to is required.")
            to_date = None
        else:
            try:
                to_date = datetime.strptime(period_to, "%Y-%m-%d")
            except ValueError:
                row_errors.append("period_to must use YYYY-MM-DD format.")
                to_date = None
        if from_date is not None and to_date is not None and to_date < from_date:
            row_errors.append("period_to cannot be earlier than period_from.")
        if promoter_username not in promoter_lookup:
            row_errors.append(f"promoter_username '{promoter_username}' must exist as an active promoter in Users sheet.")
        row_errors.extend(amount_errors)

        preview_record = {
            "period_from": period_from,
            "period_to": period_to,
            "promoter_username": promoter_username,
            "promoter_name": promoter_lookup.get(promoter_username, ""),
            "total_sales": total_sales,
        }
        if row_errors:
            errors.extend([f"Historical Sales row {index}: {message}" for message in row_errors])
            historical_rows.append(build_row_result(index, "invalid", row_errors, preview_record))
            continue

        historical_record = {
            "id": next_history_id,
            "period_from": period_from,
            "period_to": period_to,
            "promoter_username": promoter_username,
            "promoter_name": promoter_lookup[promoter_username],
            "total_sales": total_sales,
            "notes": notes,
            "updated_at": updated_at,
        }
        historical_sales.append(historical_record)
        historical_rows.append(build_row_result(index, "valid", ["Ready to import."], preview_record))
        next_history_id += 1

    if not any(normalize_text(user.get("role")).lower() == "super_admin" for user in users):
        for user in users:
            if normalize_text(user.get("username")).lower() == "admin":
                user["role"] = "super_admin"
                for row_result in user_rows:
                    if normalize_text(row_result["preview"].get("username")).lower() == "admin":
                        row_result["preview"]["role"] = "super_admin"
                        row_result["messages"] = row_result["messages"] + ["No super_admin was provided, so the admin account will be upgraded automatically."]
                break
        else:
            errors.append("Users sheet must include at least one super_admin account.")

    current = get_all_data()
    imported_data = {
        "users": users,
        "skus": skus,
        "targets": targets,
        "sales": current["sales"],
        "audit_logs": current["audit_logs"],
        "historical_sales": historical_sales,
        "correction_requests": current.get("correction_requests", []),
    }

    preview = {
        "filename": getattr(file_storage, "filename", "uploaded-workbook.xlsx"),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "is_valid": len(errors) == 0,
        "errors": errors,
        "summary": {
            "users": {"total": len(imported_users), "valid": len(users), "invalid": len(imported_users) - len(users)},
            "skus": {"total": len(imported_skus), "valid": len(skus), "invalid": len(imported_skus) - len(skus)},
            "targets": {"total": len(imported_targets), "valid": len(targets), "invalid": len(imported_targets) - len(targets)},
            "historical_sales": {"total": len(imported_historical_sales), "valid": len(historical_sales), "invalid": len(imported_historical_sales) - len(historical_sales)},
        },
        "row_results": {
            "users": user_rows,
            "skus": sku_rows,
            "targets": target_rows,
            "historical_sales": historical_rows,
        },
        "imported_data": imported_data if not errors else None,
    }
    return preview

    


def enforce_rrp_limit(
    selling_price: float,
    sku_code: str,
    sku_map: dict[str, dict[str, object]],
    errors: list[str],
    row_prefix: str = "",
) -> None:
    sku = sku_map.get(sku_code)
    if sku is None:
        return
    rrp_price = to_float(sku.get("default_price"))
    if selling_price > rrp_price:
        prefix = f"{row_prefix}: " if row_prefix else ""
        errors.append(f"{prefix}selling price cannot be more than RRP {rrp_price:.2f} for SKU {sku_code}")


def parse_sale_form(
    form: dict[str, str],
    sku_map: dict[str, dict[str, object]],
    enforce_rrp: bool = False,
) -> tuple[dict[str, object], list[str]]:
    errors: list[str] = []

    sale_date = normalize_text(form.get("sale_date"))
    location = normalize_text(form.get("location"))
    sku_code = normalize_text(form.get("sku_code"))
    quantity_raw = normalize_text(form.get("quantity"))
    selling_price_raw = normalize_text(form.get("selling_price"))
    notes = normalize_text(form.get("notes"))

    if not sale_date:
        errors.append("Sale date is required.")
    else:
        try:
            datetime.strptime(sale_date, "%Y-%m-%d")
        except ValueError:
            errors.append("Sale date must use YYYY-MM-DD format.")

    if not location:
        errors.append("Location is required.")

    if sku_code not in sku_map:
        errors.append("Please select a valid SKU.")

    quantity = 0
    try:
        quantity = int(quantity_raw)
        if quantity <= 0:
            errors.append("Quantity must be greater than zero.")
    except ValueError:
        errors.append("Quantity must be a whole number.")

    selling_price = 0.0
    try:
        selling_price = float(selling_price_raw)
        if selling_price < 0:
            errors.append("Selling price cannot be negative.")
    except ValueError:
        errors.append("Selling price must be a number.")

    sku = sku_map.get(sku_code, {"sku_name": "", "category": ""})
    if enforce_rrp and sku_code in sku_map:
        enforce_rrp_limit(selling_price, sku_code, sku_map, errors)
    sale_amount = round(quantity * selling_price, 2)

    return (
        {
            "sale_date": sale_date,
            "location": location,
            "sku_code": sku_code,
            "sku_name": sku["sku_name"],
            "category": sku["category"],
            "quantity": quantity,
            "selling_price": selling_price,
            "sale_amount": sale_amount,
            "notes": notes,
        },
        errors,
    )


def parse_correction_request_form(
    form: dict[str, str],
    sku_map: dict[str, dict[str, object]],
    enforce_rrp: bool = False,
) -> tuple[dict[str, object], list[str]]:
    sale_data, errors = parse_sale_form(form, sku_map, enforce_rrp=enforce_rrp)
    request_reason = normalize_text(form.get("request_reason"))
    if not request_reason:
        errors.append("Correction reason is required.")
    sale_data["request_reason"] = request_reason
    return sale_data, errors


def build_sale_row(
    sku_code: str = "",
    quantity: object = "",
    selling_price: object = "",
    category: str = "",
) -> dict[str, object]:
    quantity_value = "" if quantity in ("", None) else to_int(quantity)
    price_value = "" if selling_price in ("", None) else f"{to_float(selling_price):.2f}"
    total_value = 0.0
    if quantity_value != "" and price_value != "":
        total_value = to_int(quantity_value) * to_float(price_value)

    return {
        "sku_code": sku_code,
        "category": category,
        "quantity": quantity_value,
        "selling_price": price_value,
        "sale_amount": f"{total_value:.2f}" if quantity_value != "" and price_value != "" else "0.00",
    }


def blank_sale_rows(count: int = 5) -> list[dict[str, object]]:
    return [build_sale_row() for _ in range(count)]


def parse_sale_rows(
    form,
    sku_map: dict[str, dict[str, object]],
    enforce_rrp: bool = False,
) -> tuple[list[dict[str, object]], list[str], str, str]:
    errors: list[str] = []
    sale_date = normalize_text(form.get("sale_date"))
    location = normalize_text(form.get("location"))
    notes = normalize_text(form.get("notes"))

    if not sale_date:
        errors.append("Sale date is required.")
    else:
        try:
            datetime.strptime(sale_date, "%Y-%m-%d")
        except ValueError:
            errors.append("Sale date must use YYYY-MM-DD format.")

    if not location:
        errors.append("Location is required.")

    sku_codes = form.getlist("sku_code")
    quantities = form.getlist("quantity")
    selling_prices = form.getlist("selling_price")

    sale_rows: list[dict[str, object]] = []
    has_any_row = False

    max_rows = max(len(sku_codes), len(quantities), len(selling_prices), 0)
    for index in range(max_rows):
        sku_code = normalize_text(sku_codes[index] if index < len(sku_codes) else "").upper()
        quantity_raw = normalize_text(quantities[index] if index < len(quantities) else "")
        selling_price_raw = normalize_text(selling_prices[index] if index < len(selling_prices) else "")

        if not sku_code and not quantity_raw and not selling_price_raw:
            continue

        has_any_row = True
        row_errors: list[str] = []
        if sku_code not in sku_map:
            row_errors.append("select a valid SKU")

        quantity = 0
        try:
            quantity = int(quantity_raw)
            if quantity <= 0:
                row_errors.append("enter quantity greater than zero")
        except ValueError:
            row_errors.append("enter valid quantity")

        selling_price = 0.0
        try:
            selling_price = float(selling_price_raw)
            if selling_price < 0:
                row_errors.append("amount cannot be negative")
        except ValueError:
            row_errors.append("enter valid amount")

        sku = sku_map.get(sku_code, {"sku_name": "", "category": ""})
        if enforce_rrp and sku_code in sku_map:
            enforce_rrp_limit(selling_price, sku_code, sku_map, row_errors)
        if row_errors:
            errors.append(f"Row {index + 1}: " + ", ".join(row_errors) + ".")
            continue

        sale_rows.append(
            {
                "sale_date": sale_date,
                "location": location,
                "sku_code": sku_code,
                "sku_name": sku["sku_name"],
                "category": sku["category"],
                "quantity": quantity,
                "selling_price": round(selling_price, 2),
                "sale_amount": round(quantity * selling_price, 2),
                "notes": notes,
            }
        )

    if not has_any_row:
        errors.append("Enter at least one sale row before saving.")

    return sale_rows, errors, sale_date, location


def parse_target_form(form: dict[str, str], valid_promoters: set[str]) -> tuple[dict[str, object], list[str]]:
    errors: list[str] = []
    target_from = normalize_date_value(form.get("target_from"))
    target_to = normalize_date_value(form.get("target_to"))
    promoter_username = normalize_text(form.get("promoter_username")).lower()
    target_amount_raw = normalize_text(form.get("target_amount"))
    notes = normalize_text(form.get("notes"))

    if not target_from:
        errors.append("Target from date is required.")
    else:
        try:
            datetime.strptime(target_from, "%Y-%m-%d")
        except ValueError:
            errors.append("Target from date must use YYYY-MM-DD format.")

    if not target_to:
        errors.append("Target to date is required.")
    else:
        try:
            datetime.strptime(target_to, "%Y-%m-%d")
        except ValueError:
            errors.append("Target to date must use YYYY-MM-DD format.")

    if target_from and target_to:
        try:
            from_date = datetime.strptime(target_from, "%Y-%m-%d")
            to_date = datetime.strptime(target_to, "%Y-%m-%d")
            if to_date < from_date:
                errors.append("Target to date cannot be earlier than target from date.")
        except ValueError:
            pass

    if promoter_username not in valid_promoters:
        errors.append("Please select a valid promoter.")

    target_amount = 0.0
    try:
        target_amount = float(target_amount_raw)
        if target_amount < 0:
            errors.append("Target amount cannot be negative.")
    except ValueError:
        errors.append("Target amount must be a number.")

    return (
        {
            "target_from": target_from,
            "target_to": target_to,
            "promoter_username": promoter_username,
            "target_amount": round(target_amount, 2),
            "notes": notes,
        },
        errors,
    )


def parse_historical_sales_form(form: dict[str, str], valid_promoters: set[str]) -> tuple[dict[str, object], list[str]]:
    errors: list[str] = []
    period_from = normalize_date_value(form.get("period_from"))
    period_to = normalize_date_value(form.get("period_to"))
    promoter_username = normalize_text(form.get("promoter_username")).lower()
    total_sales_raw = normalize_text(form.get("total_sales"))
    notes = normalize_text(form.get("notes"))

    if not period_from:
        errors.append("Period from date is required.")
    else:
        try:
            datetime.strptime(period_from, "%Y-%m-%d")
        except ValueError:
            errors.append("Period from date must use YYYY-MM-DD format.")

    if not period_to:
        errors.append("Period to date is required.")
    else:
        try:
            datetime.strptime(period_to, "%Y-%m-%d")
        except ValueError:
            errors.append("Period to date must use YYYY-MM-DD format.")

    if period_from and period_to:
        try:
            from_date = datetime.strptime(period_from, "%Y-%m-%d")
            to_date = datetime.strptime(period_to, "%Y-%m-%d")
            if to_date < from_date:
                errors.append("Period to date cannot be earlier than period from date.")
        except ValueError:
            pass

    if promoter_username not in valid_promoters:
        errors.append("Please select a valid promoter.")

    total_sales = 0.0
    try:
        total_sales = float(total_sales_raw)
        if total_sales < 0:
            errors.append("Total sales cannot be negative.")
    except ValueError:
        errors.append("Total sales must be a number.")

    return (
        {
            "period_from": period_from,
            "period_to": period_to,
            "promoter_username": promoter_username,
            "total_sales": round(total_sales, 2),
            "notes": notes,
        },
        errors,
    )


def can_edit_sale(user: dict[str, str], sale: dict[str, object]) -> bool:
    if is_admin_role(user["role"]):
        return True
    return (
        normalize_text(sale.get("username")) == user["username"]
        and normalize_text(sale.get("sale_date")) == date.today().isoformat()
    )


def can_export_sale(user: dict[str, str], sale: dict[str, object]) -> bool:
    if is_admin_role(user["role"]):
        return True
    return normalize_text(sale.get("username")) == user["username"]


def format_sale_record(sale: dict[str, object]) -> dict[str, object]:
    return {
        **sale,
        "id": to_int(sale.get("id")),
        "quantity": to_int(sale.get("quantity")),
        "selling_price": to_float(sale.get("selling_price")),
        "sale_amount": to_float(sale.get("sale_amount")),
        "is_editable_today": normalize_text(sale.get("sale_date")) == date.today().isoformat(),
    }


def parse_user_form(
    form: dict[str, str],
    existing_usernames: set[str],
    allowed_roles: set[str],
    editing_username: str | None = None,
):
    errors: list[str] = []
    username = normalize_text(form.get("username")).lower()
    full_name = normalize_text(form.get("full_name"))
    role = normalize_text(form.get("role")).lower()
    location = normalize_text(form.get("location"))
    password = normalize_text(form.get("password"))
    active = "yes" if normalize_text(form.get("active")).lower() == "yes" else "no"

    if not username:
        errors.append("Username is required.")
    if username in existing_usernames and username != editing_username:
        errors.append("Username already exists.")
    if not full_name:
        errors.append("Full name is required.")
    if role not in allowed_roles:
        errors.append("Selected role is not allowed for your access level.")
    if not location:
        errors.append("Location is required.")
    if editing_username is None and not password:
        errors.append("Password is required for new users.")

    return {
        "username": username,
        "full_name": full_name,
        "role": role,
        "location": location,
        "password": password,
        "active": active,
    }, errors


def parse_sku_form(form: dict[str, str], existing_codes: set[str], editing_code: str | None = None):
    errors: list[str] = []
    sku_code = normalize_text(form.get("sku_code")).upper()
    sku_name = normalize_text(form.get("sku_name"))
    category = normalize_text(form.get("category"))
    default_price_raw = normalize_text(form.get("default_price"))
    active = "yes" if normalize_text(form.get("active")).lower() == "yes" else "no"

    if not sku_code:
        errors.append("SKU code is required.")
    if sku_code in existing_codes and sku_code != editing_code:
        errors.append("SKU code already exists.")
    if not sku_name:
        errors.append("SKU name is required.")
    if not category:
        errors.append("Category is required.")

    default_price = 0.0
    try:
        default_price = float(default_price_raw)
        if default_price < 0:
            errors.append("Default price cannot be negative.")
    except ValueError:
        errors.append("Default price must be a number.")

    return {
        "sku_code": sku_code,
        "sku_name": sku_name,
        "category": category,
        "default_price": round(default_price, 2),
        "active": active,
    }, errors


def build_admin_metrics(
    sales: list[dict[str, object]],
    targets: list[dict[str, object]],
    date_from: str,
    date_to: str,
    historical_sales: list[dict[str, object]] | None = None,
):
    filtered_sales = [
        format_sale_record(sale)
        for sale in sales
        if date_from <= normalize_text(sale.get("sale_date")) <= date_to
    ]
    historical_sales = historical_sales or []
    filtered_targets = [
        {
            **target,
            "target_amount": to_float(target.get("target_amount")),
            "promoter_username": normalize_text(target.get("promoter_username")).lower(),
            "promoter_name": normalize_text(target.get("promoter_name")),
            "target_from": normalize_text(target.get("target_from")),
            "target_to": normalize_text(target.get("target_to")),
        }
        for target in targets
        if normalize_text(target.get("target_from")) <= date_to
        and normalize_text(target.get("target_to")) >= date_from
        and normalize_text(target.get("promoter_username"))
    ]

    category_map: dict[str, dict[str, object]] = {}
    for sale in filtered_sales:
        category = normalize_text(sale.get("category")) or "Uncategorized"
        row = category_map.setdefault(
            category,
            {"category": category, "quantity": 0, "sales": 0.0, "target": 0.0, "achievement": 0.0},
        )
        row["quantity"] += to_int(sale.get("quantity"))
        row["sales"] += to_float(sale.get("sale_amount"))

    category_rows = []
    for row in category_map.values():
        row["target"] = 0.0
        row["achievement"] = 0.0
        category_rows.append(row)

    category_rows.sort(key=lambda row: (-to_float(row["sales"]), row["category"].lower()))

    promoter_map: dict[str, dict[str, object]] = {}
    for sale in filtered_sales:
        promoter_username = normalize_text(sale.get("username")).lower() or "unknown"
        promoter_name = normalize_text(sale.get("promoter_name")) or promoter_username or "Unknown"
        row = promoter_map.setdefault(
            promoter_username,
            {"promoter_name": promoter_name, "entries": 0, "quantity": 0, "sales": 0.0},
        )
        row["entries"] += 1
        row["quantity"] += to_int(sale.get("quantity"))
        row["sales"] += to_float(sale.get("sale_amount"))

    for target in filtered_targets:
        promoter_username = target["promoter_username"] or "unknown"
        promoter_name = target["promoter_name"] or promoter_username
        row = promoter_map.setdefault(
            promoter_username,
            {"promoter_name": promoter_name, "entries": 0, "quantity": 0, "sales": 0.0},
        )
        row["target"] = to_float(row.get("target", 0.0)) + to_float(target.get("target_amount"))

    for historical_row in historical_sales:
        period_from = normalize_text(historical_row.get("period_from"))
        period_to = normalize_text(historical_row.get("period_to"))
        if not period_from or not period_to:
            continue
        if period_from < date_from or period_to > date_to:
            continue
        promoter_username = normalize_text(historical_row.get("promoter_username")).lower() or "unknown"
        promoter_name = normalize_text(historical_row.get("promoter_name")) or promoter_username
        row = promoter_map.setdefault(
            promoter_username,
            {"promoter_name": promoter_name, "entries": 0, "quantity": 0, "sales": 0.0},
        )
        row["sales"] += to_float(historical_row.get("total_sales"))
        row["history_sources"] = to_int(row.get("history_sources")) + 1

    promoter_rows = []
    for row in promoter_map.values():
        target = to_float(row.get("target", 0.0))
        sales_amount = to_float(row.get("sales", 0.0))
        row["target"] = target
        row["achievement"] = round((sales_amount / target) * 100, 2) if target > 0 else 0.0
        promoter_rows.append(row)

    promoter_rows.sort(key=lambda row: (-to_float(row["sales"]), row["promoter_name"].lower()))
    recent_sales = sorted(filtered_sales, key=lambda sale: normalize_text(sale.get("created_at")), reverse=True)[:12]

    total_sales = sum(to_float(row.get("sales")) for row in promoter_rows)
    total_target = sum(to_float(target.get("target_amount")) for target in filtered_targets)
    return {
        "summary": {
            "total_sales": total_sales,
            "total_target": total_target,
            "achievement": round((total_sales / total_target) * 100, 2) if total_target > 0 else 0.0,
            "quantity": sum(to_int(sale.get("quantity")) for sale in filtered_sales),
            "entries": len(filtered_sales),
        },
        "category_rows": category_rows,
        "promoter_rows": promoter_rows,
        "recent_sales": recent_sales,
    }


def build_promoter_metrics(
    username: str,
    sales: list[dict[str, object]],
    targets: list[dict[str, object]],
    historical_sales: list[dict[str, object]],
    date_from: str,
    date_to: str,
) -> dict[str, object]:
    range_sales = [
        format_sale_record(sale)
        for sale in sales
        if normalize_text(sale.get("username")).lower() == username.lower()
        and date_from <= normalize_text(sale.get("sale_date")) <= date_to
    ]
    range_sales.sort(key=lambda sale: normalize_text(sale.get("created_at")), reverse=True)

    active_targets = [
        {
            **target,
            "target_amount": to_float(target.get("target_amount")),
            "target_from": normalize_text(target.get("target_from")),
            "target_to": normalize_text(target.get("target_to")),
        }
        for target in targets
        if normalize_text(target.get("promoter_username")).lower() == username.lower()
        and normalize_text(target.get("target_from")) <= date_to
        and normalize_text(target.get("target_to")) >= date_from
    ]

    historical_rows = [
        {
            **row,
            "total_sales": to_float(row.get("total_sales")),
            "period_from": normalize_text(row.get("period_from")),
            "period_to": normalize_text(row.get("period_to")),
        }
        for row in historical_sales
        if normalize_text(row.get("promoter_username")).lower() == username.lower()
        and normalize_text(row.get("period_from")) >= date_from
        and normalize_text(row.get("period_to")) <= date_to
    ]
    historical_total = sum(to_float(row.get("total_sales")) for row in historical_rows)
    live_total = sum(to_float(sale.get("sale_amount")) for sale in range_sales)

    summary = {
        "entries": len(range_sales),
        "quantity": sum(to_int(sale.get("quantity")) for sale in range_sales),
        "sales": round(live_total + historical_total, 2),
        "live_sales": round(live_total, 2),
        "historical_sales": round(historical_total, 2),
        "history_sources": len(historical_rows),
    }

    if not active_targets:
        return {
            "sales": range_sales,
            "summary": summary,
            "target_summary": {
                "has_target": False,
                "target_amount": 0.0,
                "achieved_sales": round(live_total + historical_total, 2),
                "achievement": 0.0,
                "period_from": date_from,
                "period_to": date_to,
                "target_rows": [],
            },
        }

    period_from = min(target["target_from"] for target in active_targets)
    period_to = max(target["target_to"] for target in active_targets)
    period_live_sales = [
        sale
        for sale in sales
        if normalize_text(sale.get("username")).lower() == username.lower()
        and period_from <= normalize_text(sale.get("sale_date")) <= period_to
    ]
    period_historical_total = sum(
        to_float(row.get("total_sales"))
        for row in historical_sales
        if normalize_text(row.get("promoter_username")).lower() == username.lower()
        and period_from <= normalize_text(row.get("period_from")) <= period_to
        and period_from <= normalize_text(row.get("period_to")) <= period_to
    )
    target_amount = sum(to_float(target.get("target_amount")) for target in active_targets)
    achieved_sales = sum(to_float(sale.get("sale_amount")) for sale in period_live_sales) + period_historical_total

    return {
        "sales": range_sales,
        "summary": summary,
        "target_summary": {
            "has_target": True,
            "target_amount": round(target_amount, 2),
            "achieved_sales": round(achieved_sales, 2),
            "achievement": round((achieved_sales / target_amount) * 100, 2) if target_amount > 0 else 0.0,
            "period_from": period_from,
            "period_to": period_to,
            "target_rows": active_targets,
        },
    }


def build_history_metrics(
    sales: list[dict[str, object]],
    targets: list[dict[str, object]],
    historical_sales: list[dict[str, object]],
    date_from: str,
    date_to: str,
) -> dict[str, object]:
    metrics = build_admin_metrics(sales, targets, date_from, date_to)
    promoter_map = {
        normalize_text(row.get("promoter_name")).lower(): {
            **row,
            "history_sources": 0,
        }
        for row in metrics["promoter_rows"]
    }

    for record in historical_sales:
        period_from = normalize_text(record.get("period_from"))
        period_to = normalize_text(record.get("period_to"))
        if not period_from or not period_to:
            continue
        if period_from < date_from or period_to > date_to:
            continue
        promoter_name = normalize_text(record.get("promoter_name")) or normalize_text(record.get("promoter_username")) or "Unknown"
        key = promoter_name.lower()
        row = promoter_map.setdefault(
            key,
            {
                "promoter_name": promoter_name,
                "entries": 0,
                "quantity": 0,
                "sales": 0.0,
                "target": 0.0,
                "achievement": 0.0,
                "history_sources": 0,
            },
        )
        row["sales"] += to_float(record.get("total_sales"))
        row["history_sources"] += 1

    promoter_rows = []
    for key, row in promoter_map.items():
        target = to_float(row.get("target", 0.0))
        sales_amount = to_float(row.get("sales", 0.0))
        row["achievement"] = round((sales_amount / target) * 100, 2) if target > 0 else 0.0
        promoter_rows.append(row)

    promoter_rows.sort(key=lambda row: row["promoter_name"].lower())
    total_sales = sum(to_float(row.get("sales")) for row in promoter_rows)
    total_target = sum(to_float(row.get("target")) for row in promoter_rows)

    return {
        "summary": {
            "total_sales": round(total_sales, 2),
            "total_target": round(total_target, 2),
            "achievement": round((total_sales / total_target) * 100, 2) if total_target > 0 else 0.0,
        },
        "promoter_rows": promoter_rows,
    }

    if not active_targets:
        return {
            "sales": selected_sales,
            "summary": summary,
            "target_summary": {
                "has_target": False,
                "target_amount": 0.0,
                "achieved_sales": 0.0,
                "achievement": 0.0,
                "period_from": selected_date,
                "period_to": selected_date,
                "target_rows": [],
            },
        }

    period_from = min(target["target_from"] for target in active_targets)
    period_to = max(target["target_to"] for target in active_targets)
    period_sales = [
        sale
        for sale in sales
        if normalize_text(sale.get("username")).lower() == username.lower()
        and period_from <= normalize_text(sale.get("sale_date")) <= period_to
    ]
    target_amount = sum(to_float(target.get("target_amount")) for target in active_targets)
    achieved_sales = sum(to_float(sale.get("sale_amount")) for sale in period_sales)

    return {
        "sales": selected_sales,
        "summary": summary,
        "target_summary": {
            "has_target": True,
            "target_amount": round(target_amount, 2),
            "achieved_sales": round(achieved_sales, 2),
            "achievement": round((achieved_sales / target_amount) * 100, 2) if target_amount > 0 else 0.0,
            "period_from": period_from,
            "period_to": period_to,
            "target_rows": active_targets,
        },
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user() is not None:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = normalize_text(request.form.get("username"))
        password = normalize_text(request.form.get("password"))
        user = get_user(username)

        if user is None or user["active"] != "yes" or not check_password_hash(user["password_hash"], password):
            flash("Invalid username or password.", "error")
            return render_template("login.html")

        session.clear()
        session["username"] = user["username"]
        session["full_name"] = user["full_name"]
        session["role"] = user["role"]
        session["location"] = user["location"]
        flash(f"Welcome, {user['full_name']}.", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    if is_admin_role(user["role"]):
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("promoter_dashboard"))


@app.route("/promoter")
@login_required
def promoter_dashboard():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    today = date.today()
    period = normalize_text(request.args.get("period")).lower() or "month"
    if period == "week":
        start_date = today.fromordinal(today.toordinal() - today.weekday())
        default_from = start_date.isoformat()
        default_to = today.isoformat()
    elif period == "custom":
        default_from = request.args.get("date_from", today.replace(day=1).isoformat())
        default_to = request.args.get("date_to", today.isoformat())
    else:
        period = "month"
        default_from = today.replace(day=1).isoformat()
        default_to = today.isoformat()

    date_from = normalize_date_value(request.args.get("date_from")) or default_from
    date_to = normalize_date_value(request.args.get("date_to")) or default_to
    data = get_all_data()
    pending_sale_ids = {
        to_int(request["sale_id"])
        for request in data.get("correction_requests", [])
        if normalize_text(request.get("requested_by")) == user["username"]
        and normalize_text(request.get("status")).lower() == "pending"
    }
    metrics = build_promoter_metrics(
        user["username"],
        data["sales"],
        data["targets"],
        data.get("historical_sales", []),
        date_from,
        date_to,
    )
    sales = metrics["sales"]
    for sale in sales:
        sale["has_pending_correction"] = sale["id"] in pending_sale_ids

    return render_template(
        "promoter_dashboard.html",
        sales=sales,
        summary=metrics["summary"],
        target_summary=metrics["target_summary"],
        period=period,
        date_from=date_from,
        date_to=date_to,
        today_iso=today.isoformat(),
    )


@app.route("/corrections")
@login_required
def my_corrections():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    requests = [
        {
            **request_row,
            "id": to_int(request_row.get("id")),
            "sale_id": to_int(request_row.get("sale_id")),
            "requested_quantity": to_int(request_row.get("requested_quantity")),
            "requested_selling_price": to_float(request_row.get("requested_selling_price")),
            "requested_sale_amount": to_float(request_row.get("requested_sale_amount")),
        }
        for request_row in get_all_data()["correction_requests"]
        if normalize_text(request_row.get("requested_by")) == user["username"]
    ]
    requests.sort(key=lambda row: normalize_text(row.get("created_at")), reverse=True)
    return render_template("correction_requests.html", requests=requests, admin_mode=False)


@app.route("/admin")
@admin_required
def admin_dashboard():
    today = date.today()
    date_from = request.args.get("date_from", today.replace(day=1).isoformat())
    date_to = request.args.get("date_to", today.isoformat())

    data = get_all_data()
    metrics = build_admin_metrics(data["sales"], data["targets"], date_from, date_to, data.get("historical_sales", []))
    categories = sorted({normalize_text(sku.get("category")) for sku in data["skus"] if normalize_text(sku.get("category"))})

    return render_template(
        "admin_dashboard.html",
        date_from=date_from,
        date_to=date_to,
        summary=metrics["summary"],
        category_rows=metrics["category_rows"],
        promoter_rows=metrics["promoter_rows"],
        recent_sales=metrics["recent_sales"],
        categories=categories,
    )


@app.route("/admin/history", methods=["GET", "POST"])
@admin_required
def admin_history():
    data = get_all_data()
    promoters = sorted(
        [
            {
                "username": normalize_text(user.get("username")).lower(),
                "full_name": normalize_text(user.get("full_name")),
                "location": normalize_text(user.get("location")),
            }
            for user in data["users"]
            if normalize_text(user.get("role")).lower() == "promoter" and normalize_text(user.get("active")).lower() == "yes"
        ],
        key=lambda item: item["full_name"].lower(),
    )
    promoter_usernames = {promoter["username"] for promoter in promoters}
    editing_history_id = to_int(request.args.get("edit"))
    editing_history = next((row for row in data.get("historical_sales", []) if to_int(row.get("id")) == editing_history_id), None) if editing_history_id else None

    if request.method == "POST":
        form_mode = normalize_text(request.form.get("form_mode")) or "create"
        editing_history_id = to_int(request.form.get("editing_history_id")) if form_mode == "edit" else 0
        payload, errors = parse_historical_sales_form(request.form, promoter_usernames)
        if not errors:
            duplicate_history = next(
                (
                    row
                    for row in data.get("historical_sales", [])
                    if normalize_text(row.get("period_from")) == payload["period_from"]
                    and normalize_text(row.get("period_to")) == payload["period_to"]
                    and normalize_text(row.get("promoter_username")).lower() == payload["promoter_username"]
                    and (form_mode != "edit" or to_int(row.get("id")) != editing_history_id)
                ),
                None,
            )
            if duplicate_history is not None:
                errors.append("A historical sales record already exists for this promoter and period.")

        if not errors:
            promoter = next((item for item in promoters if item["username"] == payload["promoter_username"]), None)
            if form_mode == "edit" and editing_history_id:
                record = next((item for item in data.get("historical_sales", []) if to_int(item.get("id")) == editing_history_id), None)
                if record is None:
                    flash("Historical sales record not found.", "error")
                    return redirect(url_for("admin_history"))
                old_snapshot = (
                    f"{normalize_text(record.get('period_from'))}:{normalize_text(record.get('period_to'))}:"
                    f"{normalize_text(record.get('promoter_username'))}:{to_float(record.get('total_sales')):.2f}"
                )
                record["period_from"] = payload["period_from"]
                record["period_to"] = payload["period_to"]
                record["promoter_username"] = payload["promoter_username"]
                record["promoter_name"] = promoter["full_name"] if promoter else payload["promoter_username"]
                record["total_sales"] = payload["total_sales"]
                record["notes"] = payload["notes"]
                record["updated_at"] = datetime.now().isoformat(timespec="seconds")
                append_audit_log(
                    data,
                    "update",
                    "historical_sale",
                    editing_history_id,
                    f"Updated historical sale {old_snapshot} -> {payload['period_from']}:{payload['period_to']}:{payload['promoter_username']}:{payload['total_sales']:.2f}",
                )
                flash("Historical sales record updated.", "success")
            else:
                data.setdefault("historical_sales", []).append(
                    {
                        "id": next_id(data.get("historical_sales", [])),
                        "period_from": payload["period_from"],
                        "period_to": payload["period_to"],
                        "promoter_username": payload["promoter_username"],
                        "promoter_name": promoter["full_name"] if promoter else payload["promoter_username"],
                        "total_sales": payload["total_sales"],
                        "notes": payload["notes"],
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                )
                append_audit_log(
                    data,
                    "create",
                    "historical_sale",
                    f"{payload['period_from']}:{payload['period_to']}:{payload['promoter_username']}",
                    f"Historical sales set to {payload['total_sales']:.2f}",
                )
                flash("Historical sales record saved.", "success")
            write_all_data(data)
            return redirect(url_for("admin_history"))

        for error in errors:
            flash(error, "error")

    today = date.today()
    period = normalize_text(request.args.get("period")).lower() or "month"
    if period == "week":
        start_date = today.fromordinal(today.toordinal() - today.weekday())
        default_from = start_date.isoformat()
        default_to = today.isoformat()
    elif period == "custom":
        default_from = request.args.get("date_from", today.replace(day=1).isoformat())
        default_to = request.args.get("date_to", today.isoformat())
    else:
        period = "month"
        default_from = today.replace(day=1).isoformat()
        default_to = today.isoformat()

    date_from = normalize_text(request.args.get("date_from")) or default_from
    date_to = normalize_text(request.args.get("date_to")) or default_to

    metrics = build_history_metrics(data["sales"], data["targets"], data.get("historical_sales", []), date_from, date_to)
    history_rows = sorted(
        [
            {
                **row,
                "id": to_int(row.get("id")),
                "total_sales": to_float(row.get("total_sales")),
            }
            for row in data.get("historical_sales", [])
        ],
        key=lambda row: (normalize_text(row.get("period_from")), normalize_text(row.get("promoter_name"))),
        reverse=True,
    )
    history_form = {
        "form_mode": "edit" if editing_history else "create",
        "editing_history_id": to_int(editing_history.get("id")) if editing_history else 0,
        "period_from": normalize_text(editing_history.get("period_from")) if editing_history else "",
        "period_to": normalize_text(editing_history.get("period_to")) if editing_history else "",
        "promoter_username": normalize_text(editing_history.get("promoter_username")).lower() if editing_history else "",
        "total_sales": to_float(editing_history.get("total_sales")) if editing_history else "",
        "notes": normalize_text(editing_history.get("notes")) if editing_history else "",
    }

    return render_template(
        "admin_history.html",
        period=period,
        date_from=date_from,
        date_to=date_to,
        summary=metrics["summary"],
        promoter_rows=metrics["promoter_rows"],
        promoters=promoters,
        history_form=history_form,
        editing_history=editing_history,
        history_rows=history_rows,
    )


@app.route("/sales/new", methods=["GET", "POST"])
@login_required
def create_sale():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    sku_map = get_sku_map()
    skus = list(sku_map.values())
    fixed_sale_date = date.today().isoformat()
    fixed_location = user["location"]

    if request.method == "POST":
        sale_rows, errors, sale_date, location = parse_sale_rows(
            request.form,
            sku_map,
            enforce_rrp=not is_admin_role(user["role"]),
        )
        sale_date = fixed_sale_date
        location = fixed_location
        for row in sale_rows:
            row["sale_date"] = fixed_sale_date
            row["location"] = fixed_location
        if not errors:
            data = get_all_data()
            timestamp = datetime.now().isoformat(timespec="seconds")
            next_sale_id = next_id(data["sales"])
            for row in sale_rows:
                new_sale = {
                    "id": next_sale_id,
                    "sale_date": row["sale_date"],
                    "username": user["username"],
                    "promoter_name": user["full_name"],
                    "location": row["location"],
                    "sku_code": row["sku_code"],
                    "sku_name": row["sku_name"],
                    "category": row["category"],
                    "quantity": row["quantity"],
                    "selling_price": row["selling_price"],
                    "sale_amount": row["sale_amount"],
                    "notes": row["notes"],
                    "created_at": timestamp,
                    "updated_at": timestamp,
                }
                data["sales"].append(new_sale)
                append_audit_log(
                    data,
                    "create",
                    "sale",
                    new_sale["id"],
                    f"Created sale for {new_sale['sku_code']} qty {new_sale['quantity']} on {new_sale['sale_date']}",
                )
                next_sale_id += 1
            write_all_data(data)
            flash(f"{len(sale_rows)} sale row(s) saved successfully.", "success")
            return redirect(url_for("promoter_dashboard", sale_date=sale_date))

        for error in errors:
            flash(error, "error")
        sku_codes = request.form.getlist("sku_code")
        quantities = request.form.getlist("quantity")
        selling_prices = request.form.getlist("selling_price")
        sale_rows_view = []
        max_rows = max(len(sku_codes), len(quantities), len(selling_prices), 5)
        for index in range(max_rows):
            sku_code = normalize_text(sku_codes[index] if index < len(sku_codes) else "").upper()
            sku = sku_map.get(sku_code, {"category": ""})
            quantity = quantities[index] if index < len(quantities) else ""
            selling_price = selling_prices[index] if index < len(selling_prices) else ""
            sale_rows_view.append(build_sale_row(sku_code, quantity, selling_price, sku.get("category", "")))
        sale = {
            "sale_date": fixed_sale_date,
            "location": fixed_location,
            "notes": normalize_text(request.form.get("notes")),
        }
    else:
        sale = {"sale_date": fixed_sale_date, "location": fixed_location, "notes": ""}
        sale_rows_view = blank_sale_rows()

    return render_template(
        "sale_form.html",
        form_title="Update Daily Sale",
        submit_label="Save & Add Next Bunch",
        sale=sale,
        sale_rows=sale_rows_view,
        skus=skus,
        user=user,
        is_bulk_entry=True,
    )


@app.route("/sales/<int:sale_id>/edit", methods=["GET", "POST"])
@login_required
def edit_sale(sale_id: int):
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    data = get_all_data()
    sale = next((format_sale_record(record) for record in data["sales"] if to_int(record.get("id")) == sale_id), None)
    if sale is None or not can_edit_sale(user, sale):
        flash("Sale record not found or access denied.", "error")
        return redirect(url_for("dashboard"))

    sku_map = get_sku_map()
    skus = list(sku_map.values())

    if request.method == "POST":
        sale_data, errors = parse_sale_form(
            request.form,
            sku_map,
            enforce_rrp=not is_admin_role(user["role"]),
        )
        if not errors:
            for record in data["sales"]:
                if to_int(record.get("id")) != sale_id:
                    continue
                old_snapshot = (
                    f"old={normalize_text(record.get('sku_code'))}/{normalize_text(record.get('sale_date'))}/"
                    f"{to_int(record.get('quantity'))}/{to_float(record.get('selling_price')):.2f}"
                )
                record["sku_code"] = sale_data["sku_code"]
                record["sku_name"] = sale_data["sku_name"]
                record["category"] = sale_data["category"]
                record["quantity"] = sale_data["quantity"]
                record["selling_price"] = sale_data["selling_price"]
                record["sale_amount"] = sale_data["sale_amount"]
                record["notes"] = sale_data["notes"]
                record["updated_at"] = datetime.now().isoformat(timespec="seconds")
                append_audit_log(
                    data,
                    "update",
                    "sale",
                    sale_id,
                    f"{old_snapshot} new={record['sku_code']}/{record['sale_date']}/{record['quantity']}/{to_float(record['selling_price']):.2f}",
                )
                break
            write_all_data(data)
            flash("Sale record updated.", "success")
            return redirect(url_for("promoter_dashboard", sale_date=sale_data["sale_date"]))

        for error in errors:
            flash(error, "error")
        sale = {**sale, **dict(request.form)}

    return render_template(
        "sale_form.html",
        form_title="Edit Sale",
        submit_label="Update Sale",
        sale=sale,
        sale_rows=[build_sale_row(sale["sku_code"], sale["quantity"], sale["selling_price"], sale["category"])],
        skus=skus,
        user=user,
        is_bulk_entry=False,
    )


@app.route("/sales/<int:sale_id>/request-correction", methods=["GET", "POST"])
@login_required
def request_correction(sale_id: int):
    user = current_user()
    if user is None:
        return redirect(url_for("login"))
    if is_admin_role(user["role"]):
        flash("Admins can edit sales directly. Correction requests are for promoters.", "error")
        return redirect(url_for("dashboard"))

    data = get_all_data()
    sale = next((format_sale_record(record) for record in data["sales"] if to_int(record.get("id")) == sale_id), None)
    if sale is None or normalize_text(sale.get("username")) != user["username"]:
        flash("Sale record not found or access denied.", "error")
        return redirect(url_for("promoter_dashboard"))
    if sale["is_editable_today"]:
        flash("Today's sales can be edited directly. A correction request is only needed for locked sales.", "error")
        return redirect(url_for("promoter_dashboard", sale_date=sale["sale_date"]))

    existing_pending = next(
        (
            row
            for row in data.get("correction_requests", [])
            if to_int(row.get("sale_id")) == sale_id
            and normalize_text(row.get("requested_by")) == user["username"]
            and normalize_text(row.get("status")).lower() == "pending"
        ),
        None,
    )
    if existing_pending is not None:
        flash("A pending correction request already exists for this sale.", "error")
        return redirect(url_for("my_corrections"))

    sku_map = get_sku_map()
    skus = list(sku_map.values())

    if request.method == "POST":
        request_data, errors = parse_correction_request_form(
            request.form,
            sku_map,
            enforce_rrp=not is_admin_role(user["role"]),
        )
        if not errors:
            timestamp = datetime.now().isoformat(timespec="seconds")
            correction = {
                "id": next_id(data.get("correction_requests", [])),
                "sale_id": sale_id,
                "sale_date": sale["sale_date"],
                "requested_by": user["username"],
                "promoter_name": sale["promoter_name"],
                "location": sale["location"],
                "current_sku_code": sale["sku_code"],
                "current_sku_name": sale["sku_name"],
                "current_category": sale["category"],
                "current_quantity": sale["quantity"],
                "current_selling_price": sale["selling_price"],
                "current_sale_amount": sale["sale_amount"],
                "requested_sku_code": request_data["sku_code"],
                "requested_sku_name": request_data["sku_name"],
                "requested_category": request_data["category"],
                "requested_quantity": request_data["quantity"],
                "requested_selling_price": request_data["selling_price"],
                "requested_sale_amount": request_data["sale_amount"],
                "request_reason": request_data["request_reason"],
                "status": "pending",
                "reviewer_username": "",
                "review_notes": "",
                "created_at": timestamp,
                "reviewed_at": "",
            }
            data.setdefault("correction_requests", []).append(correction)
            append_audit_log(
                data,
                "request",
                "correction",
                correction["id"],
                f"Requested correction for sale {sale_id}",
            )
            write_all_data(data)
            flash("Correction request submitted for admin review.", "success")
            return redirect(url_for("my_corrections"))

        for error in errors:
            flash(error, "error")
        requested_sale = {**sale, **dict(request.form)}
    else:
        requested_sale = {
            "sale_date": sale["sale_date"],
            "location": sale["location"],
            "sku_code": sale["sku_code"],
            "category": sale["category"],
            "quantity": sale["quantity"],
            "selling_price": sale["selling_price"],
            "sale_amount": sale["sale_amount"],
            "request_reason": "",
        }

    return render_template(
        "correction_request_form.html",
        sale=sale,
        requested_sale=requested_sale,
        skus=skus,
        user=user,
    )


@app.post("/sales/<int:sale_id>/delete")
@login_required
def delete_sale(sale_id: int):
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    data = get_all_data()
    sale = next((record for record in data["sales"] if to_int(record.get("id")) == sale_id), None)
    if sale is None or not can_edit_sale(user, sale):
        flash("Sale record not found or access denied.", "error")
        return redirect(url_for("dashboard"))

    append_audit_log(
        data,
        "delete",
        "sale",
        sale_id,
        f"Deleted sale for {normalize_text(sale.get('sku_code'))} on {normalize_text(sale.get('sale_date'))}",
    )
    data["sales"] = [record for record in data["sales"] if to_int(record.get("id")) != sale_id]
    write_all_data(data)
    flash("Sale record deleted.", "success")
    return redirect(url_for("promoter_dashboard", sale_date=normalize_text(sale.get("sale_date"))))


@app.get("/sales/export")
@login_required
def export_sales():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    selected_date = normalize_date_value(request.args.get("sale_date"))
    date_from = normalize_date_value(request.args.get("date_from"))
    date_to = normalize_date_value(request.args.get("date_to"))
    if not date_from and not date_to:
        date_from = selected_date or date.today().isoformat()
        date_to = selected_date or date.today().isoformat()
    else:
        date_from = date_from or date.today().isoformat()
        date_to = date_to or date_from

    sales = [
        format_sale_record(sale)
        for sale in get_all_data()["sales"]
        if date_from <= normalize_text(sale.get("sale_date")) <= date_to and can_export_sale(user, sale)
    ]
    sales.sort(key=lambda sale: normalize_text(sale.get("created_at")), reverse=True)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Sale Date", "Promoter", "Location", "SKU Code", "SKU Name", "Category", "Quantity", "Selling Price", "Sale Amount", "Notes"])
    for sale in sales:
        writer.writerow(
            [
                sale["sale_date"],
                sale["promoter_name"],
                sale["location"],
                sale["sku_code"],
                sale["sku_name"],
                sale["category"],
                sale["quantity"],
                f"{sale['selling_price']:.2f}",
                f"{sale['sale_amount']:.2f}",
                normalize_text(sale.get("notes")),
            ]
        )

    filename_suffix = f"{date_from}_to_{date_to}"
    filename = f"{user['username']}_sales_{filename_suffix}.csv" if not is_admin_role(user["role"]) else f"all_sales_{filename_suffix}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/admin/corrections")
@admin_required
def admin_corrections():
    requests = [
        {
            **request_row,
            "id": to_int(request_row.get("id")),
            "sale_id": to_int(request_row.get("sale_id")),
            "current_quantity": to_int(request_row.get("current_quantity")),
            "requested_quantity": to_int(request_row.get("requested_quantity")),
            "current_selling_price": to_float(request_row.get("current_selling_price")),
            "requested_selling_price": to_float(request_row.get("requested_selling_price")),
            "current_sale_amount": to_float(request_row.get("current_sale_amount")),
            "requested_sale_amount": to_float(request_row.get("requested_sale_amount")),
        }
        for request_row in get_all_data().get("correction_requests", [])
    ]
    requests.sort(
        key=lambda row: (
            0 if normalize_text(row.get("status")).lower() == "pending" else 1,
            normalize_text(row.get("created_at")),
        ),
        reverse=False,
    )
    return render_template("correction_requests.html", requests=requests, admin_mode=True)


@app.post("/admin/corrections/<int:request_id>/approve")
@admin_required
def approve_correction(request_id: int):
    reviewer = current_user()
    if reviewer is None:
        return redirect(url_for("login"))
    data = get_all_data()
    request_row = next((row for row in data.get("correction_requests", []) if to_int(row.get("id")) == request_id), None)
    if request_row is None:
        flash("Correction request not found.", "error")
        return redirect(url_for("admin_corrections"))
    if normalize_text(request_row.get("status")).lower() != "pending":
        flash("This correction request has already been reviewed.", "error")
        return redirect(url_for("admin_corrections"))

    sale = next((row for row in data["sales"] if to_int(row.get("id")) == to_int(request_row.get("sale_id"))), None)
    if sale is None:
        flash("Original sale not found.", "error")
        return redirect(url_for("admin_corrections"))

    sale["sku_code"] = request_row.get("requested_sku_code")
    sale["sku_name"] = request_row.get("requested_sku_name")
    sale["category"] = request_row.get("requested_category")
    sale["quantity"] = request_row.get("requested_quantity")
    sale["selling_price"] = request_row.get("requested_selling_price")
    sale["sale_amount"] = request_row.get("requested_sale_amount")
    sale["notes"] = f"{normalize_text(sale.get('notes'))} | correction approved".strip(" |")
    sale["updated_at"] = datetime.now().isoformat(timespec="seconds")

    request_row["status"] = "approved"
    request_row["reviewer_username"] = reviewer["username"]
    request_row["review_notes"] = normalize_text(request.form.get("review_notes")) or "Approved"
    request_row["reviewed_at"] = datetime.now().isoformat(timespec="seconds")
    append_audit_log(data, "approve", "correction", request_id, f"Approved correction request {request_id}")
    write_all_data(data)
    flash("Correction request approved and sale updated.", "success")
    return redirect(url_for("admin_corrections"))


@app.post("/admin/corrections/<int:request_id>/reject")
@admin_required
def reject_correction(request_id: int):
    reviewer = current_user()
    if reviewer is None:
        return redirect(url_for("login"))
    data = get_all_data()
    request_row = next((row for row in data.get("correction_requests", []) if to_int(row.get("id")) == request_id), None)
    if request_row is None:
        flash("Correction request not found.", "error")
        return redirect(url_for("admin_corrections"))
    if normalize_text(request_row.get("status")).lower() != "pending":
        flash("This correction request has already been reviewed.", "error")
        return redirect(url_for("admin_corrections"))

    request_row["status"] = "rejected"
    request_row["reviewer_username"] = reviewer["username"]
    request_row["review_notes"] = normalize_text(request.form.get("review_notes")) or "Rejected"
    request_row["reviewed_at"] = datetime.now().isoformat(timespec="seconds")
    append_audit_log(data, "reject", "correction", request_id, f"Rejected correction request {request_id}")
    write_all_data(data)
    flash("Correction request rejected.", "success")
    return redirect(url_for("admin_corrections"))


@app.route("/admin/targets", methods=["GET", "POST"])
@admin_required
def manage_targets():
    data = get_all_data()
    editing_target_id = to_int(request.args.get("edit"))
    editing_target = next((target for target in data["targets"] if to_int(target.get("id")) == editing_target_id), None) if editing_target_id else None
    promoters = sorted(
        [
            {
                "username": normalize_text(user.get("username")).lower(),
                "full_name": normalize_text(user.get("full_name")),
                "location": normalize_text(user.get("location")),
            }
            for user in data["users"]
            if normalize_text(user.get("role")).lower() == "promoter" and normalize_text(user.get("active")).lower() == "yes"
        ],
        key=lambda item: item["full_name"].lower(),
    )
    promoter_usernames = {promoter["username"] for promoter in promoters}

    if request.method == "POST":
        form_mode = normalize_text(request.form.get("form_mode")) or "create"
        editing_target_id = to_int(request.form.get("editing_target_id")) if form_mode == "edit" else 0
        target_data, errors = parse_target_form(request.form, promoter_usernames)
        if not errors:
            duplicate_target = next(
                (
                    target
                    for target in data["targets"]
                    if normalize_text(target.get("target_from")) == target_data["target_from"]
                    and normalize_text(target.get("target_to")) == target_data["target_to"]
                    and normalize_text(target.get("promoter_username")).lower() == target_data["promoter_username"]
                    and (form_mode != "edit" or to_int(target.get("id")) != editing_target_id)
                ),
                None,
            )
            if duplicate_target is not None:
                errors.append("A target already exists for this promoter and period.")

        if not errors:
            promoter = next((item for item in promoters if item["username"] == target_data["promoter_username"]), None)
            if form_mode == "edit" and editing_target_id:
                target = next((item for item in data["targets"] if to_int(item.get("id")) == editing_target_id), None)
                if target is None:
                    flash("Target record not found.", "error")
                    return redirect(url_for("manage_targets"))
                old_snapshot = (
                    f"{normalize_text(target.get('target_from'))}:{normalize_text(target.get('target_to'))}:"
                    f"{normalize_text(target.get('promoter_username'))}:{to_float(target.get('target_amount')):.2f}"
                )
                target["promoter_name"] = promoter["full_name"] if promoter else target.get("promoter_name", "")
                target["promoter_username"] = target_data["promoter_username"]
                target["target_from"] = target_data["target_from"]
                target["target_to"] = target_data["target_to"]
                target["target_amount"] = target_data["target_amount"]
                target["notes"] = target_data["notes"]
                target["updated_at"] = datetime.now().isoformat(timespec="seconds")
                append_audit_log(
                    data,
                    "update",
                    "target",
                    editing_target_id,
                    f"Updated target {old_snapshot} -> {target_data['target_from']}:{target_data['target_to']}:{target_data['promoter_username']}:{target_data['target_amount']:.2f}",
                )
                flash("Promoter target updated.", "success")
            else:
                promoter = next((item for item in promoters if item["username"] == target_data["promoter_username"]), None)
                data["targets"].append(
                    {
                        "id": next_id(data["targets"]),
                        "target_from": target_data["target_from"],
                        "target_to": target_data["target_to"],
                        "promoter_username": target_data["promoter_username"],
                        "promoter_name": promoter["full_name"] if promoter else target_data["promoter_username"],
                        "target_amount": target_data["target_amount"],
                        "notes": target_data["notes"],
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                )
                append_audit_log(
                    data,
                    "save",
                    "target",
                    f"{target_data['target_from']}:{target_data['target_to']}:{target_data['promoter_username']}",
                    f"Target set to {target_data['target_amount']:.2f}",
                )
            write_all_data(data)
            return redirect(url_for("manage_targets"))

        for error in errors:
            flash(error, "error")

    targets = sorted(
        [
            {
                **target,
                "id": to_int(target.get("id")),
                "target_amount": to_float(target.get("target_amount")),
            }
            for target in data["targets"]
        ],
        key=lambda item: (normalize_text(item.get("target_from")), normalize_text(item.get("promoter_name"))),
        reverse=True,
    )

    target_form = {
        "form_mode": "edit" if editing_target else "create",
        "editing_target_id": to_int(editing_target.get("id")) if editing_target else 0,
        "target_from": normalize_text(editing_target.get("target_from")) if editing_target else "",
        "target_to": normalize_text(editing_target.get("target_to")) if editing_target else "",
        "promoter_username": normalize_text(editing_target.get("promoter_username")).lower() if editing_target else "",
        "target_amount": to_float(editing_target.get("target_amount")) if editing_target else "",
        "notes": normalize_text(editing_target.get("notes")) if editing_target else "",
    }

    return render_template("targets.html", promoters=promoters, targets=targets, target_form=target_form, editing_target=editing_target)


@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def manage_users():
    actor = current_user()
    if actor is None:
        return redirect(url_for("login"))

    data = get_all_data()
    editing_username = normalize_text(request.args.get("edit")).lower() or None
    editing_user = None
    if editing_username:
        editing_user = next((user for user in data["users"] if normalize_text(user.get("username")).lower() == editing_username), None)

    if editing_user and actor["role"] != "super_admin" and normalize_text(editing_user.get("role")) != "promoter":
        flash("Only super admin can manage admin accounts.", "error")
        return redirect(url_for("manage_users"))

    if request.method == "POST":
        form_mode = normalize_text(request.form.get("form_mode")) or "create"
        editing_username = normalize_text(request.form.get("editing_username")).lower() or None
        allowed_roles = {"promoter"} if actor["role"] == "admin" else {"promoter", "admin", "super_admin"}
        payload, errors = parse_user_form(
            request.form,
            {normalize_text(user.get("username")).lower() for user in data["users"]},
            allowed_roles,
            editing_username if form_mode == "edit" else None,
        )
        if not errors:
            if actor["role"] != "super_admin" and payload["role"] != "promoter":
                flash("Only super admin can create or update admin accounts.", "error")
                return redirect(url_for("manage_users"))
            if form_mode == "edit" and editing_username:
                for user in data["users"]:
                    if normalize_text(user.get("username")).lower() != editing_username:
                        continue
                    if actor["role"] != "super_admin" and normalize_text(user.get("role")) != "promoter":
                        flash("Only super admin can update admin accounts.", "error")
                        return redirect(url_for("manage_users"))
                    user["full_name"] = payload["full_name"]
                    user["role"] = payload["role"]
                    user["location"] = payload["location"]
                    user["active"] = payload["active"]
                    if payload["password"]:
                        user["password_hash"] = generate_password_hash(payload["password"])
                    append_audit_log(data, "update", "user", editing_username, f"Updated user {editing_username}")
                    break
            else:
                data["users"].append(
                    {
                        "username": payload["username"],
                        "password_hash": generate_password_hash(payload["password"]),
                        "full_name": payload["full_name"],
                        "role": payload["role"],
                        "location": payload["location"],
                        "active": payload["active"],
                    }
                )
                append_audit_log(data, "create", "user", payload["username"], f"Created user {payload['username']}")

            write_all_data(data)
            flash("User saved successfully.", "success")
            return redirect(url_for("manage_users"))

        for error in errors:
            flash(error, "error")
        editing_user = {**(editing_user or {}), **dict(request.form)}

    users = sorted(data["users"], key=lambda item: normalize_text(item.get("username")).lower())
    if actor["role"] != "super_admin":
        users = [user for user in users if normalize_text(user.get("role")) == "promoter"]
    return render_template(
        "users.html",
        users=users,
        editing_user=editing_user,
        can_manage_admins=(actor["role"] == "super_admin"),
        current_role=actor["role"],
    )


@app.route("/admin/skus", methods=["GET", "POST"])
@admin_required
def manage_skus():
    data = get_all_data()
    editing_code = normalize_text(request.args.get("edit")).upper() or None
    editing_sku = None
    if editing_code:
        editing_sku = next((sku for sku in data["skus"] if normalize_text(sku.get("sku_code")).upper() == editing_code), None)

    if request.method == "POST":
        form_mode = normalize_text(request.form.get("form_mode")) or "create"
        editing_code = normalize_text(request.form.get("editing_code")).upper() or None
        payload, errors = parse_sku_form(
            request.form,
            {normalize_text(sku.get("sku_code")).upper() for sku in data["skus"]},
            editing_code if form_mode == "edit" else None,
        )
        if not errors:
            if form_mode == "edit" and editing_code:
                for sku in data["skus"]:
                    if normalize_text(sku.get("sku_code")).upper() != editing_code:
                        continue
                    sku["sku_name"] = payload["sku_name"]
                    sku["category"] = payload["category"]
                    sku["default_price"] = payload["default_price"]
                    sku["active"] = payload["active"]
                    append_audit_log(data, "update", "sku", editing_code, f"Updated SKU {editing_code}")
                    break
            else:
                data["skus"].append(payload)
                append_audit_log(data, "create", "sku", payload["sku_code"], f"Created SKU {payload['sku_code']}")

            write_all_data(data)
            flash("SKU saved successfully.", "success")
            return redirect(url_for("manage_skus"))

        for error in errors:
            flash(error, "error")
        editing_sku = {**(editing_sku or {}), **dict(request.form)}

    skus = sorted(data["skus"], key=lambda item: normalize_text(item.get("sku_code")).upper())
    return render_template("skus.html", skus=skus, editing_sku=editing_sku)


@app.get("/admin/audit")
@admin_required
def view_audit_log():
    audit_logs = [
        {
            **log,
            "id": to_int(log.get("id")),
        }
        for log in get_all_data()["audit_logs"]
    ]
    audit_logs.sort(key=lambda item: normalize_text(item.get("event_time")), reverse=True)
    return render_template("audit.html", audit_logs=audit_logs[:200])


@app.get("/admin/setup")
@admin_required
def admin_setup():
    actor = current_user()
    if actor is None:
        return redirect(url_for("login"))
    data = get_all_data()
    active_users = sum(1 for user in data["users"] if normalize_text(user.get("active")).lower() == "yes")
    active_skus = sum(1 for sku in data["skus"] if normalize_text(sku.get("active")).lower() == "yes")
    target_count = len(data["targets"])
    sales_count = len(data["sales"])
    return render_template(
        "admin_setup.html",
        active_users=active_users,
        active_skus=active_skus,
        target_count=target_count,
        sales_count=sales_count,
        can_manage_imports=(actor["role"] == "super_admin"),
        can_manage_admins=(actor["role"] == "super_admin"),
    )


@app.route("/admin/import", methods=["GET", "POST"])
@super_admin_required
def import_setup():
    actor = current_user()
    preview = load_import_preview(app.config["DATABASE"], actor["username"]) if actor is not None else None

    if request.method == "POST":
        uploaded_file = request.files.get("backend_file")
        if uploaded_file is None or not uploaded_file.filename:
            flash("Please choose an Excel file to import.", "error")
            return render_template("import_setup.html", preview=preview)

        preview = import_backend_template(uploaded_file)
        if actor is not None:
            save_import_preview(app.config["DATABASE"], actor["username"], preview)

        if preview["is_valid"]:
            flash("Preview ready. Review the rows below, then confirm the import.", "success")
        else:
            flash("Preview found validation issues. Fix the flagged rows and upload again.", "error")
        return render_template("import_setup.html", preview=preview)

    return render_template("import_setup.html", preview=preview)


@app.post("/admin/import/confirm")
@super_admin_required
def confirm_import_setup():
    actor = current_user()
    if actor is None:
        return redirect(url_for("login"))

    preview = load_import_preview(app.config["DATABASE"], actor["username"])
    if preview is None:
        flash("No import preview was found. Upload a workbook first.", "error")
        return redirect(url_for("import_setup"))
    if not preview.get("is_valid"):
        flash("This preview still has validation issues. Upload a corrected workbook first.", "error")
        return redirect(url_for("import_setup"))

    imported_data = preview.get("imported_data")
    if not isinstance(imported_data, dict):
        flash("Import preview is incomplete. Upload the workbook again.", "error")
        return redirect(url_for("import_setup"))

    append_audit_log(
        imported_data,
        "import",
        "backend_template",
        preview.get("filename", "uploaded-workbook.xlsx"),
        "Imported users, SKUs, and targets from confirmed workbook preview",
    )
    write_all_data(imported_data)
    clear_import_preview(app.config["DATABASE"], actor["username"])
    flash("Backend template imported successfully. Plain passwords were converted to secure hashes.", "success")
    return redirect(url_for("manage_users"))


@app.post("/admin/import/discard")
@super_admin_required
def discard_import_setup():
    actor = current_user()
    if actor is not None:
        clear_import_preview(app.config["DATABASE"], actor["username"])
    flash("Import preview cleared.", "success")
    return redirect(url_for("import_setup"))


@app.post("/admin/setup/clear-sales")
@admin_required
def clear_sales():
    data = get_all_data()
    data["sales"] = []
    append_audit_log(
        data,
        "clear",
        "sales",
        "all",
        "Cleared all sales records from admin setup hub",
    )
    write_all_data(data)
    flash("All sales records were cleared successfully.", "success")
    return redirect(url_for("admin_setup"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=False)
