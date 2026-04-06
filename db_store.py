from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover - optional for local SQLite-only runs
    psycopg = None
    dict_row = None


USERS_COLUMNS = ["username", "password_hash", "full_name", "role", "location", "active"]
SKUS_COLUMNS = ["sku_code", "sku_name", "category", "default_price", "active"]
TARGETS_COLUMNS = ["id", "target_from", "target_to", "promoter_username", "promoter_name", "target_amount", "notes", "updated_at"]
SALES_COLUMNS = [
    "id",
    "sale_date",
    "username",
    "promoter_name",
    "location",
    "sku_code",
    "sku_name",
    "category",
    "quantity",
    "selling_price",
    "sale_amount",
    "notes",
    "created_at",
    "updated_at",
]
AUDIT_COLUMNS = ["id", "event_time", "actor_username", "actor_role", "action", "entity_type", "entity_id", "details"]
HISTORICAL_SALES_COLUMNS = ["id", "period_from", "period_to", "promoter_username", "promoter_name", "total_sales", "notes", "updated_at"]
CORRECTION_REQUEST_COLUMNS = [
    "id",
    "sale_id",
    "sale_date",
    "requested_by",
    "promoter_name",
    "location",
    "current_sku_code",
    "current_sku_name",
    "current_category",
    "current_quantity",
    "current_selling_price",
    "current_sale_amount",
    "requested_sku_code",
    "requested_sku_name",
    "requested_category",
    "requested_quantity",
    "requested_selling_price",
    "requested_sale_amount",
    "request_reason",
    "status",
    "reviewer_username",
    "review_notes",
    "created_at",
    "reviewed_at",
]


DEFAULT_USERS = [
    {
        "username": "admin",
        "password_hash": generate_password_hash("admin123"),
        "full_name": "Portal Admin",
        "role": "super_admin",
        "location": "Head Office",
        "active": "yes",
    },
    {
        "username": "promoter1",
        "password_hash": generate_password_hash("promoter123"),
        "full_name": "Promoter One",
        "role": "promoter",
        "location": "Dubai Mall",
        "active": "yes",
    },
]


DEFAULT_SKUS = [
    {
        "sku_code": "A1001",
        "sku_name": "Anker Charger 20W",
        "category": "Chargers",
        "default_price": 79.0,
        "active": "yes",
    },
    {
        "sku_code": "A2001",
        "sku_name": "Anker Power Bank 10000mAh",
        "category": "Power Banks",
        "default_price": 149.0,
        "active": "yes",
    },
    {
        "sku_code": "A3001",
        "sku_name": "Anker USB-C Cable",
        "category": "Cables",
        "default_price": 39.0,
        "active": "yes",
    },
]

SCHEMA_STATEMENTS = [
    """
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY,
        password_hash TEXT NOT NULL,
        full_name TEXT NOT NULL,
        role TEXT NOT NULL,
        location TEXT NOT NULL,
        active TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS skus (
        sku_code TEXT PRIMARY KEY,
        sku_name TEXT NOT NULL,
        category TEXT NOT NULL,
        default_price REAL NOT NULL,
        active TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS targets (
        id INTEGER PRIMARY KEY,
        target_from TEXT NOT NULL,
        target_to TEXT NOT NULL,
        promoter_username TEXT NOT NULL,
        promoter_name TEXT NOT NULL,
        target_amount REAL NOT NULL,
        notes TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY,
        sale_date TEXT NOT NULL,
        username TEXT NOT NULL,
        promoter_name TEXT NOT NULL,
        location TEXT NOT NULL,
        sku_code TEXT NOT NULL,
        sku_name TEXT NOT NULL,
        category TEXT NOT NULL,
        quantity INTEGER NOT NULL,
        selling_price REAL NOT NULL,
        sale_amount REAL NOT NULL,
        notes TEXT NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS audit_logs (
        id INTEGER PRIMARY KEY,
        event_time TEXT NOT NULL,
        actor_username TEXT NOT NULL,
        actor_role TEXT NOT NULL,
        action TEXT NOT NULL,
        entity_type TEXT NOT NULL,
        entity_id TEXT NOT NULL,
        details TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS historical_sales (
        id INTEGER PRIMARY KEY,
        period_from TEXT NOT NULL,
        period_to TEXT NOT NULL,
        promoter_username TEXT NOT NULL,
        promoter_name TEXT NOT NULL,
        total_sales REAL NOT NULL,
        notes TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS correction_requests (
        id INTEGER PRIMARY KEY,
        sale_id INTEGER NOT NULL,
        sale_date TEXT NOT NULL,
        requested_by TEXT NOT NULL,
        promoter_name TEXT NOT NULL,
        location TEXT NOT NULL,
        current_sku_code TEXT NOT NULL,
        current_sku_name TEXT NOT NULL,
        current_category TEXT NOT NULL,
        current_quantity INTEGER NOT NULL,
        current_selling_price REAL NOT NULL,
        current_sale_amount REAL NOT NULL,
        requested_sku_code TEXT NOT NULL,
        requested_sku_name TEXT NOT NULL,
        requested_category TEXT NOT NULL,
        requested_quantity INTEGER NOT NULL,
        requested_selling_price REAL NOT NULL,
        requested_sale_amount REAL NOT NULL,
        request_reason TEXT NOT NULL,
        status TEXT NOT NULL,
        reviewer_username TEXT NOT NULL,
        review_notes TEXT NOT NULL,
        created_at TEXT NOT NULL,
        reviewed_at TEXT NOT NULL
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS import_previews (
        username TEXT PRIMARY KEY,
        preview_json TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """,
]


def normalize_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def to_int(value: object) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def to_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def is_postgres_target(target: str | Path) -> bool:
    text = str(target)
    return text.startswith("postgres://") or text.startswith("postgresql://")


def is_sqlite_connection(conn: Any) -> bool:
    return isinstance(conn, sqlite3.Connection)


def connect(db_path: str | Path):
    target = str(db_path)
    if is_postgres_target(target):
        if psycopg is None:
            raise RuntimeError("psycopg is required when DATABASE_URL points to PostgreSQL.")
        return psycopg.connect(target, row_factory=dict_row)

    conn = sqlite3.connect(target)
    conn.row_factory = sqlite3.Row
    return conn


def placeholders(conn: Any, count: int) -> str:
    token = "?" if is_sqlite_connection(conn) else "%s"
    return ", ".join([token] * count)


def create_schema(conn: Any) -> None:
    for statement in SCHEMA_STATEMENTS:
        conn.execute(statement)
    conn.commit()


def table_count(conn: Any, table_name: str) -> int:
    return int(conn.execute(f"SELECT COUNT(*) AS count FROM {table_name}").fetchone()["count"])


def read_sheet(workbook, sheet_name: str) -> list[dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [normalize_text(cell) for cell in rows[0]]
    records: list[dict[str, object]] = []
    for row in rows[1:]:
        if row is None or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record = {}
        for index, column in enumerate(header):
            record[column] = "" if index >= len(row) or row[index] is None else row[index]
        records.append(record)
    return records


def migrate_from_workbook(conn: Any, workbook_path: str | Path) -> None:
    path = Path(workbook_path)
    if not path.exists():
        seed_defaults(conn)
        return

    wb = load_workbook(path, data_only=True)

    users_raw = read_sheet(wb, "Users")
    skus_raw = read_sheet(wb, "SKUs")
    targets_raw = read_sheet(wb, "Targets")
    sales_raw = read_sheet(wb, "Sales")
    audit_raw = read_sheet(wb, "Audit Log")
    wb.close()

    users: list[dict[str, object]] = []
    for row in users_raw:
        username = normalize_text(row.get("username")).lower()
        password_hash = normalize_text(row.get("password_hash"))
        plain_password = normalize_text(row.get("plain_password"))
        if plain_password:
            password_hash = generate_password_hash(plain_password)
        if not username or not password_hash:
            continue
        users.append(
            {
                "username": username,
                "password_hash": password_hash,
                "full_name": normalize_text(row.get("full_name")),
                "role": normalize_text(row.get("role")) or "promoter",
                "location": normalize_text(row.get("location")),
                "active": normalize_text(row.get("active")) or "yes",
            }
        )

    if users and not any(normalize_text(user["role"]).lower() == "super_admin" for user in users):
        for user in users:
            if user["username"] == "sabith":
                user["role"] = "super_admin"
                break
        else:
            for user in users:
                if user["username"] == "admin":
                    user["role"] = "super_admin"
                    break

    skus = [
        {
            "sku_code": normalize_text(row.get("sku_code")).upper(),
            "sku_name": normalize_text(row.get("sku_name")),
            "category": normalize_text(row.get("category")),
            "default_price": to_float(row.get("default_price")),
            "active": normalize_text(row.get("active")) or "yes",
        }
        for row in skus_raw
        if normalize_text(row.get("sku_code"))
    ]

    promoter_lookup = {user["username"]: user["full_name"] for user in users if normalize_text(user["role"]).lower() == "promoter"}
    targets = []
    for row in targets_raw:
        promoter_username = normalize_text(row.get("promoter_username") or row.get("username")).lower()
        target_from = normalize_text(row.get("target_from") or row.get("target_date"))
        target_to = normalize_text(row.get("target_to") or row.get("target_date"))
        if not promoter_username or not target_from or not target_to:
            continue
        targets.append(
            {
                "id": to_int(row.get("id")) or (len(targets) + 1),
                "target_from": target_from,
                "target_to": target_to,
                "promoter_username": promoter_username,
                "promoter_name": normalize_text(row.get("promoter_name")) or promoter_lookup.get(promoter_username, promoter_username),
                "target_amount": to_float(row.get("target_amount")),
                "notes": normalize_text(row.get("notes")),
                "updated_at": normalize_text(row.get("updated_at")),
            }
        )

    sales = []
    for row in sales_raw:
        sale_id = to_int(row.get("id"))
        if sale_id == 0:
            continue
        sales.append(
            {
                "id": sale_id,
                "sale_date": normalize_text(row.get("sale_date")),
                "username": normalize_text(row.get("username")).lower(),
                "promoter_name": normalize_text(row.get("promoter_name")),
                "location": normalize_text(row.get("location")),
                "sku_code": normalize_text(row.get("sku_code")).upper(),
                "sku_name": normalize_text(row.get("sku_name")),
                "category": normalize_text(row.get("category")),
                "quantity": to_int(row.get("quantity")),
                "selling_price": to_float(row.get("selling_price")),
                "sale_amount": to_float(row.get("sale_amount")),
                "notes": normalize_text(row.get("notes")),
                "created_at": normalize_text(row.get("created_at")),
                "updated_at": normalize_text(row.get("updated_at")),
            }
        )

    audit_logs = []
    for row in audit_raw:
        log_id = to_int(row.get("id"))
        if log_id == 0:
            continue
        audit_logs.append(
            {
                "id": log_id,
                "event_time": normalize_text(row.get("event_time")),
                "actor_username": normalize_text(row.get("actor_username")),
                "actor_role": normalize_text(row.get("actor_role")),
                "action": normalize_text(row.get("action")),
                "entity_type": normalize_text(row.get("entity_type")),
                "entity_id": normalize_text(row.get("entity_id")),
                "details": normalize_text(row.get("details")),
            }
        )

    if not users:
        seed_defaults(conn)
        return

    save_all_data(
        conn,
        {
            "users": users,
            "skus": skus or DEFAULT_SKUS,
            "targets": targets,
            "sales": sales,
            "audit_logs": audit_logs,
            "historical_sales": [],
            "correction_requests": [],
        },
    )


def seed_defaults(conn: Any) -> None:
    save_all_data(
        conn,
        {
            "users": DEFAULT_USERS,
            "skus": DEFAULT_SKUS,
            "targets": [],
            "sales": [],
            "audit_logs": [],
            "historical_sales": [],
            "correction_requests": [],
        },
    )


def ensure_database(db_path: str | Path, workbook_path: str | Path) -> None:
    conn = connect(db_path)
    create_schema(conn)
    if table_count(conn, "users") == 0:
        migrate_from_workbook(conn, workbook_path)
    conn.close()


def rows_to_dicts(cursor_rows: list[Any]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row in cursor_rows:
        rows.append(dict(row))
    return rows


def load_all_data(db_path: str | Path) -> dict[str, list[dict[str, object]]]:
    conn = connect(db_path)
    data = {
        "users": rows_to_dicts(conn.execute("SELECT username, password_hash, full_name, role, location, active FROM users ORDER BY username").fetchall()),
        "skus": rows_to_dicts(conn.execute("SELECT sku_code, sku_name, category, default_price, active FROM skus ORDER BY sku_code").fetchall()),
        "targets": rows_to_dicts(conn.execute("SELECT id, target_from, target_to, promoter_username, promoter_name, target_amount, notes, updated_at FROM targets ORDER BY id").fetchall()),
        "sales": rows_to_dicts(conn.execute("SELECT id, sale_date, username, promoter_name, location, sku_code, sku_name, category, quantity, selling_price, sale_amount, notes, created_at, updated_at FROM sales ORDER BY id").fetchall()),
        "audit_logs": rows_to_dicts(conn.execute("SELECT id, event_time, actor_username, actor_role, action, entity_type, entity_id, details FROM audit_logs ORDER BY id").fetchall()),
        "historical_sales": rows_to_dicts(conn.execute("SELECT id, period_from, period_to, promoter_username, promoter_name, total_sales, notes, updated_at FROM historical_sales ORDER BY id").fetchall()),
        "correction_requests": rows_to_dicts(conn.execute("SELECT id, sale_id, sale_date, requested_by, promoter_name, location, current_sku_code, current_sku_name, current_category, current_quantity, current_selling_price, current_sale_amount, requested_sku_code, requested_sku_name, requested_category, requested_quantity, requested_selling_price, requested_sale_amount, request_reason, status, reviewer_username, review_notes, created_at, reviewed_at FROM correction_requests ORDER BY id").fetchall()),
    }
    conn.close()
    return data


def execute_insert_many(conn: Any, table_name: str, columns: list[str], rows: list[tuple[Any, ...]]) -> None:
    if not rows:
        return
    columns_sql = ", ".join(columns)
    sql = f"INSERT INTO {table_name} ({columns_sql}) VALUES ({placeholders(conn, len(columns))})"
    conn.executemany(sql, rows)


def save_all_data(conn_or_path: Any, data: dict[str, list[dict[str, object]]]) -> None:
    owns_conn = not hasattr(conn_or_path, "execute") or isinstance(conn_or_path, (str, Path))
    conn = connect(conn_or_path) if owns_conn else conn_or_path

    conn.execute("DELETE FROM users")
    execute_insert_many(
        conn,
        "users",
        USERS_COLUMNS,
        [(row.get("username"), row.get("password_hash"), row.get("full_name"), row.get("role"), row.get("location"), row.get("active")) for row in data["users"]],
    )

    conn.execute("DELETE FROM skus")
    execute_insert_many(
        conn,
        "skus",
        SKUS_COLUMNS,
        [(row.get("sku_code"), row.get("sku_name"), row.get("category"), row.get("default_price"), row.get("active")) for row in data["skus"]],
    )

    conn.execute("DELETE FROM targets")
    execute_insert_many(
        conn,
        "targets",
        TARGETS_COLUMNS,
        [(row.get("id"), row.get("target_from"), row.get("target_to"), row.get("promoter_username"), row.get("promoter_name"), row.get("target_amount"), row.get("notes"), row.get("updated_at")) for row in data["targets"]],
    )

    conn.execute("DELETE FROM sales")
    execute_insert_many(
        conn,
        "sales",
        SALES_COLUMNS,
        [(row.get("id"), row.get("sale_date"), row.get("username"), row.get("promoter_name"), row.get("location"), row.get("sku_code"), row.get("sku_name"), row.get("category"), row.get("quantity"), row.get("selling_price"), row.get("sale_amount"), row.get("notes"), row.get("created_at"), row.get("updated_at")) for row in data["sales"]],
    )

    conn.execute("DELETE FROM audit_logs")
    execute_insert_many(
        conn,
        "audit_logs",
        AUDIT_COLUMNS,
        [(row.get("id"), row.get("event_time"), row.get("actor_username"), row.get("actor_role"), row.get("action"), row.get("entity_type"), row.get("entity_id"), row.get("details")) for row in data["audit_logs"]],
    )

    conn.execute("DELETE FROM historical_sales")
    execute_insert_many(
        conn,
        "historical_sales",
        HISTORICAL_SALES_COLUMNS,
        [
            (
                row.get("id"),
                row.get("period_from"),
                row.get("period_to"),
                row.get("promoter_username"),
                row.get("promoter_name"),
                row.get("total_sales"),
                row.get("notes"),
                row.get("updated_at"),
            )
            for row in data.get("historical_sales", [])
        ],
    )

    conn.execute("DELETE FROM correction_requests")
    execute_insert_many(
        conn,
        "correction_requests",
        CORRECTION_REQUEST_COLUMNS,
        [
            (
                row.get("id"),
                row.get("sale_id"),
                row.get("sale_date"),
                row.get("requested_by"),
                row.get("promoter_name"),
                row.get("location"),
                row.get("current_sku_code"),
                row.get("current_sku_name"),
                row.get("current_category"),
                row.get("current_quantity"),
                row.get("current_selling_price"),
                row.get("current_sale_amount"),
                row.get("requested_sku_code"),
                row.get("requested_sku_name"),
                row.get("requested_category"),
                row.get("requested_quantity"),
                row.get("requested_selling_price"),
                row.get("requested_sale_amount"),
                row.get("request_reason"),
                row.get("status"),
                row.get("reviewer_username"),
                row.get("review_notes"),
                row.get("created_at"),
                row.get("reviewed_at"),
            )
            for row in data.get("correction_requests", [])
        ],
    )

    conn.commit()
    if owns_conn:
        conn.close()


def load_import_preview(db_path: str | Path, username: str) -> dict[str, object] | None:
    conn = connect(db_path)
    create_schema(conn)
    row = conn.execute("SELECT preview_json FROM import_previews WHERE username = ?", (username,)).fetchone() if is_sqlite_connection(conn) else conn.execute("SELECT preview_json FROM import_previews WHERE username = %s", (username,)).fetchone()
    conn.close()
    if row is None:
        return None
    try:
        return json.loads(row["preview_json"])
    except (KeyError, TypeError, json.JSONDecodeError):
        return None


def save_import_preview(db_path: str | Path, username: str, preview: dict[str, object]) -> None:
    conn = connect(db_path)
    create_schema(conn)
    payload = json.dumps(preview)
    updated_at = datetime.now(timezone.utc).isoformat()
    sql = """
    INSERT INTO import_previews (username, preview_json, updated_at)
    VALUES ({placeholder}, {placeholder}, {placeholder})
    ON CONFLICT (username) DO UPDATE
    SET preview_json = EXCLUDED.preview_json,
        updated_at = EXCLUDED.updated_at
    """
    placeholder = "?" if is_sqlite_connection(conn) else "%s"
    conn.execute(sql.format(placeholder=placeholder), (username, payload, updated_at))
    conn.commit()
    conn.close()


def clear_import_preview(db_path: str | Path, username: str) -> None:
    conn = connect(db_path)
    create_schema(conn)
    conn.execute("DELETE FROM import_previews WHERE username = ?", (username,)) if is_sqlite_connection(conn) else conn.execute("DELETE FROM import_previews WHERE username = %s", (username,))
    conn.commit()
    conn.close()
